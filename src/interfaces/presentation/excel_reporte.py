from io import BytesIO


def _openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        return Workbook, BarChart, Reference, Font, PatternFill, Border, Side, Alignment
    except ImportError as exc:
        raise RuntimeError("Falta instalar openpyxl. Ejecute: pip install -r requirements.txt") from exc


def _money(v):
    try:
        return float(v or 0)
    except Exception:
        return 0


def _fecha(c):
    return str(getattr(c, "fecha", ""))[:10]


def _estado(c):
    return str(getattr(c, "estado", "pendiente") or "pendiente").strip().lower().replace(" ", "_").replace("-", "_")


def _orden_cantidad(fila):
    return fila[1]


def _contador(cotizaciones, campo):
    datos = {}
    for c in cotizaciones:
        clave = getattr(c, campo, "") or "Sin dato"
        datos[clave] = datos.get(clave, 0) + 1
    filas = list(datos.items())
    filas.sort(key=_orden_cantidad, reverse=True)
    return filas


def _auto_ancho(ws):
    for col in ws.columns:
        letra = col[0].column_letter
        maximo = 12
        for celda in col:
            maximo = max(maximo, len(str(celda.value or "")) + 2)
        ws.column_dimensions[letra].width = min(maximo, 34)


def _tabla(ws, fila_inicio, encabezados, filas, Font, PatternFill, Border, Side, Alignment):
    rojo = PatternFill("solid", fgColor="E63946")
    borde = Border(bottom=Side(style="thin", color="DDDDDD"))
    for i, titulo in enumerate(encabezados, 1):
        celda = ws.cell(fila_inicio, i, titulo)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = rojo
        celda.alignment = Alignment(horizontal="center", vertical="center")
    for r, fila in enumerate(filas, fila_inicio + 1):
        for c, valor in enumerate(fila, 1):
            celda = ws.cell(r, c, valor)
            celda.border = borde
            celda.alignment = Alignment(vertical="center")
    _auto_ancho(ws)


def _agregar_grafico(ws, titulo, data_range, cats_range, celda, BarChart, Reference):
    chart = BarChart()
    chart.title = titulo
    chart.y_axis.title = "Cantidad"
    chart.x_axis.title = "Categoría"
    data = Reference(ws, min_col=data_range[0], min_row=data_range[1], max_row=data_range[2])
    cats = Reference(ws, min_col=cats_range[0], min_row=cats_range[1], max_row=cats_range[2])
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 14
    ws.add_chart(chart, celda)


def generar_excel_reporte(cotizaciones, ranking, filtros, por_marca, por_plan):
    Workbook, BarChart, Reference, Font, PatternFill, Border, Side, Alignment = _openpyxl()
    wb = Workbook()

    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "AutoVentas Pro - Reporte gerencial"
    ws["A1"].font = Font(size=16, bold=True, color="E63946")
    ws["A3"] = "Filtros aplicados"
    ws["B3"] = filtros or "Sin filtros"

    total = len(cotizaciones)
    ventas = [c for c in cotizaciones if _estado(c) == "cerrada"]
    cerradas = len(ventas)
    ingreso = sum(_money(getattr(c, "precio_vehiculo", 0)) for c in ventas)
    tasa = round((cerradas / total) * 100, 2) if total else 0

    resumen = [
        ["Indicador", "Valor"],
        ["Cotizaciones", total],
        ["Cerradas", cerradas],
        ["Tasa de cierre", str(tasa) + "%"],
        ["Ingreso por ventas", ingreso],
    ]
    _tabla(ws, 5, resumen[0], resumen[1:], Font, PatternFill, Border, Side, Alignment)
    ws["B9"].number_format = '$#,##0.00'

    marca_datos = list(por_marca) or [("Sin datos", 0)]
    rojo = PatternFill("solid", fgColor="E63946")
    ws["D5"] = "Marca"
    ws["E5"] = "Cotizaciones"
    for celda in [ws["D5"], ws["E5"]]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = rojo
        celda.alignment = Alignment(horizontal="center")
    for idx, (marca, cantidad) in enumerate(marca_datos, 6):
        ws.cell(idx, 4, marca)
        ws.cell(idx, 5, cantidad)
    if marca_datos and marca_datos[0][0] != "Sin datos":
        _agregar_grafico(ws, "Cotizaciones por marca", (5, 5, 5 + len(marca_datos)), (4, 6, 5 + len(marca_datos)), "G5", BarChart, Reference)
    _auto_ancho(ws)

    ws = wb.create_sheet("Cotizaciones")
    filas = []
    for c in cotizaciones:
        filas.append([
            getattr(c, "id", ""),
            _fecha(c),
            getattr(c, "nombre_cliente", ""),
            getattr(c, "vehiculo_id", ""),
            getattr(c, "plan_id", ""),
            getattr(c, "asesor_id", ""),
            _money(getattr(c, "cuota_mensual", 0)),
            _estado(c),
        ])
    _tabla(ws, 1, ["ID", "Fecha", "Cliente", "Vehículo", "Plan", "Asesor ID", "Cuota", "Estado"], filas, Font, PatternFill, Border, Side, Alignment)
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 7).number_format = '$#,##0.00'

    ws = wb.create_sheet("Ranking asesores")
    filas = []
    for r in ranking:
        filas.append([r.get("nombre", ""), r.get("cotizaciones", 0), r.get("cerradas", 0), _money(r.get("valor", 0)), r.get("rating", "")])
    _tabla(ws, 1, ["Asesor", "Cotizaciones", "Cerradas", "Valor", "Rating"], filas, Font, PatternFill, Border, Side, Alignment)
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 4).number_format = '$#,##0.00'

    ws = wb.create_sheet("Por marca")
    marca_filas = list(por_marca) or [("Sin datos", 0)]
    _tabla(ws, 1, ["Marca", "Cotizaciones"], marca_filas, Font, PatternFill, Border, Side, Alignment)
    if len(marca_filas) > 0:
        _agregar_grafico(ws, "Cotizaciones por marca", (2, 1, ws.max_row), (1, 2, ws.max_row), "D2", BarChart, Reference)

    ws = wb.create_sheet("Por plan")
    _tabla(ws, 1, ["Plan", "Cotizaciones"], list(por_plan) or [("Sin datos", 0)], Font, PatternFill, Border, Side, Alignment)

    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)
    return archivo.getvalue()


def generar_excel_cotizaciones(cotizaciones):
    Workbook, BarChart, Reference, Font, PatternFill, Border, Side, Alignment = _openpyxl()
    wb = Workbook()

    total = len(cotizaciones)
    por_estado = _contador(cotizaciones, "estado") or [("Sin datos", 0)]
    por_vehiculo = _contador(cotizaciones, "vehiculo_id") or [("Sin datos", 0)]

    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "AutoVentas Pro - Cotizaciones"
    ws["A1"].font = Font(size=16, bold=True, color="E63946")
    _tabla(ws, 3, ["Indicador", "Valor"], [["Total cotizaciones", total]], Font, PatternFill, Border, Side, Alignment)
    _tabla(ws, 7, ["Estado", "Cantidad"], por_estado, Font, PatternFill, Border, Side, Alignment)
    if por_estado:
        _agregar_grafico(ws, "Cotizaciones por estado", (2, 7, 7 + len(por_estado)), (1, 8, 7 + len(por_estado)), "D7", BarChart, Reference)

    ws = wb.create_sheet("Cotizaciones")
    filas = []
    for c in cotizaciones:
        filas.append([
            getattr(c, "id", ""), _fecha(c), getattr(c, "nombre_cliente", ""),
            getattr(c, "cedula_cliente", ""), getattr(c, "email_cliente", ""),
            getattr(c, "telefono_cliente", ""), getattr(c, "ciudad_cliente", ""),
            getattr(c, "vehiculo_id", ""), getattr(c, "plan_id", ""),
            _money(getattr(c, "entrada", 0)), getattr(c, "plazo_meses", 0),
            _money(getattr(c, "cuota_mensual", 0)), _estado(c),
        ])
    _tabla(ws, 1, ["ID", "Fecha", "Cliente", "Cédula", "Email", "Teléfono", "Ciudad", "Vehículo", "Plan", "Entrada", "Plazo", "Cuota", "Estado"], filas, Font, PatternFill, Border, Side, Alignment)
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 10).number_format = '$#,##0.00'
        ws.cell(row, 12).number_format = '$#,##0.00'

    ws = wb.create_sheet("Por vehículo")
    _tabla(ws, 1, ["Vehículo", "Cotizaciones"], por_vehiculo, Font, PatternFill, Border, Side, Alignment)
    if por_vehiculo:
        _agregar_grafico(ws, "Cotizaciones por vehículo", (2, 1, ws.max_row), (1, 2, ws.max_row), "D2", BarChart, Reference)

    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)
    return archivo.getvalue()
